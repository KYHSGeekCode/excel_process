from openpyxl import load_workbook

# 필요한 정보:
# 거래처명 <- 거래처명 : C
# 사업자번호 <- 사업자번호 D
# 거래처종류? <- ?
# 업태 <- 업태 N
# 종목 <- 종목 O
# 대표자명 <- 대표자 F
# 세무사업자주소 <- 주소 Q
# 전화번호 <- 전화번호 S
# 은행명 <- 지급은행 K
# 계좌번호 <- 지급계좌번호 L
# 예금주명 <- 예금주명 M

거래처명 = 2
사업자번호 = 3
업태 = 13
종목 = 14
대표자 = 5
주소 = 16
전화번호 = 18
지급은행 = 10
지급계좌번호 = 11
예금주명 = 12


def convert_row(one_row):
    name = one_row[거래처명].value
    num = one_row[사업자번호].value
    typee = one_row[업태].value
    kind = one_row[종목].value
    person = one_row[대표자].value
    addr = one_row[주소].value
    phone = one_row[전화번호].value
    bank = one_row[지급은행].value
    moneynumber = one_row[지급계좌번호].value
    moneyname = one_row[예금주명].value
    if num == '000-00-00000' or num is None or num == '':
        num = ''

    return [
        name,
        '',
        num,
        '',
        '',
        '영리법인, 비용거래처' if num != '' else '',
        '정상',
        '국내',
        '대한민국',
        typee,
        kind,
        person,
        addr,
        phone,
        bank,
        moneynumber,
        moneyname
    ]


def write_row(ws, rowcnt, converted):
    for i, v in enumerate(converted):
        if v is None:
            v = ''
        ws.cell(rowcnt, i + 1, v)


def main():
    # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook("거래처.xlsx", data_only=True)
    # 시트 이름으로 불러오기
    load_ws = load_wb['거래처관리- 20210824']

    # out_wb = Workbook()
    dest_filename = 'output.xlsx'
    # ws1 = out_wb.active
    # ws1.title = 'result'
    target_wb = load_workbook('거래처관리_템플릿_20210820- 입력 양식.xlsx', data_only=True)
    target_sheet_title = '1. 거래처관리'
    target_ws = target_wb[target_sheet_title]

    passcnt = 0
    rowcnt = 3
    for row in load_ws.rows:
        if passcnt < 2:
            passcnt = passcnt + 1
            continue
        converted = convert_row(row)
        write_row(target_ws, rowcnt, converted)
        rowcnt = rowcnt + 1
        print(converted)
    #
    target_wb.save(dest_filename)


if __name__ == '__main__':
    main()
